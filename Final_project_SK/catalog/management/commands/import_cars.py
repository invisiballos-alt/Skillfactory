import os
from datetime import datetime
from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from django.contrib.auth.models import User
from catalog.models import (
    Machine, VehicleModel, EngineModel, TransmissionModel,
    DriveAxleModel, SteerAxleModel, ServiceCompany
)

class Command(BaseCommand):
    help = 'Импорт машин из файла Excel data.xlsx'

    def handle(self, *args, **options):
        file_path = 'data.xlsx'
        
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f"Файл {file_path} не найден в корне проекта!"))
            return

        wb = load_workbook(file_path, data_only=True)
        sheet = wb.active 

        self.stdout.write(self.style.WARNING("Начало импорта машин..."))

        # Читаем строго с 3-й строки, как в Excel
        for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
            if not row or row[2] is None: # Если заводской номер пустой — пропускаем строку
                continue

            try:
                # 1. Извлекаем данные по точным индексам колонок Excel
                v_model_name = str(row[1]).strip()
                serial_number = str(row[2]).strip()
                engine_model_name = str(row[3]).strip()
                engine_serial_number = str(row[4]).strip()
                trans_model_name = str(row[5]).strip()
                transmission_serial_number = str(row[6]).strip()
                drive_model_name = str(row[7]).strip()
                drive_axle_serial_number = str(row[8]).strip()
                steer_model_name = str(row[9]).strip()
                steer_axle_serial_number = str(row[10]).strip()
                
                # Обработка даты отгрузки (колонка 13 в Excel -> индекс 12 в Python)
                raw_date = row[11]
                if isinstance(raw_date, datetime):
                    shipping_date = raw_date.date()
                else:
                    shipping_date = datetime.strptime(str(raw_date).strip(), "%d.%m.%Y").date()

                # Генерируем строку договора: "12423 + номер строки от дата_отгрузки"
                formatted_date_str = shipping_date.strftime("%d.%m.%Y")
                supply_contract = f"12423{row_idx} от {formatted_date_str}"

                # Поля со скриншота по индексам (13 - грузополучатель, 14 - адрес)
                client_name = str(row[12]).strip()
                consignee = str(row[13]).strip() if row[13] else ""
                operation_location = str(row[14]).strip() if row[14] else ""
                
                # Доп опции, покупатель, сервис (индексы 15, 16)
                additional_options = str(row[15]).strip() if row[15] else "Стандарт"
                service_company_name = str(row[16]).strip()

                # 2. Получаем связанные объекты из базы
                vehicle_model = VehicleModel.objects.get(name=v_model_name)
                engine_model = EngineModel.objects.get(name=engine_model_name)
                transmission_model = TransmissionModel.objects.get(name=trans_model_name)
                drive_axle_model = DriveAxleModel.objects.get(name=drive_model_name)
                steer_axle_model = SteerAxleModel.objects.get(name=steer_model_name)
                service_company = ServiceCompany.objects.get(name=service_company_name)

                client_user = User.objects.filter(first_name=client_name).first()

                # 3. Сохраняем в базу данных Machine
                machine, created = Machine.objects.update_or_create(
                    serial_number=serial_number,
                    defaults={
                        'vehicle_model': vehicle_model,
                        'engine_model': engine_model,
                        'engine_serial_number': engine_serial_number,
                        'transmission_model': transmission_model,
                        'transmission_serial_number': transmission_serial_number,
                        'drive_axle_model': drive_axle_model,
                        'drive_axle_serial_number': drive_axle_serial_number,
                        'steer_axle_model': steer_axle_model,
                        'steer_axle_serial_number': steer_axle_serial_number,
                        
                        # Ваши точные имена полей из models.py:
                        'supply_contract': supply_contract,
                        'shipping_date': shipping_date,
                        'consignee': consignee,
                        'operation_location': operation_location,
                        'additional_options': additional_options,
                        
                        'client': client_user,
                        'service_company': service_company,
                    }
                )

                if created:
                    self.stdout.write(self.style.SUCCESS(f"Строка {row_idx}: Машина №{serial_number} успешно добавлена."))
                else:
                    self.stdout.write(self.style.WARNING(f"Строка {row_idx}: Машина №{serial_number} обновлена."))

            except Exception as e:
                self.stdout.write(self.style.ERROR(f"Ошибка в строке {row_idx}: {e}"))

        self.stdout.write(self.style.SUCCESS("Импорт машин успешно завершен!"))
