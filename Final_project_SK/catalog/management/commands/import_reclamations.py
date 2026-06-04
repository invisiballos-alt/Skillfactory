import os
from datetime import datetime
from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from catalog.models import Machine, Reclamation, FailureNode, RecoveryMethod

class Command(BaseCommand):
    help = 'Импорт данных рекламаций из Excel data.xlsx (Вкладка Рекламации)'

    def handle(self, *args, **options):
        file_path = 'data.xlsx'
        
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f"Файл {file_path} не найден в корень проекта!"))
            return

        wb = load_workbook(file_path, data_only=True)
        # Открываем третью вкладку рекламаций
        sheet = wb.worksheets[2] 

        self.stdout.write(self.style.WARNING("Начало импорта рекламаций..."))

        for row_idx in range(3, sheet.max_row + 1):
            serial_number_raw = sheet[f'A{row_idx}'].value # Читаем зав. номер из колонки A
            if not serial_number_raw:
                continue

            try:
                serial_number = str(serial_number_raw).strip()
                # Убираем дробную часть, если номер определился как "3.0"
                if serial_number.endswith('.0'):
                    serial_number = serial_number[:-2]
                # Дописываем нули до 4 символов (например, "0003")
                serial_number = serial_number.zfill(4)
                
                # Дата отказа (Колонка B)
                raw_refusal_date = sheet[f'B{row_idx}'].value
                if isinstance(raw_refusal_date, datetime):
                    refusal_date = raw_refusal_date.date()
                else:
                    refusal_date = datetime.strptime(str(raw_refusal_date).strip(), "%d.%m.%Y").date()

                operating_hours = int(sheet[f'C{row_idx}'].value)
                node_name = str(sheet[f'D{row_idx}'].value).strip()
                failure_description = str(sheet[f'E{row_idx}'].value).strip()
                method_name = str(sheet[f'F{row_idx}'].value).strip()
                
                # Читаем запчасти из колонки G и проверяем её же на пустоту
                spare_parts = str(sheet[f'G{row_idx}'].value).strip() if sheet[f'G{row_idx}'].value else ""
                
                # Дата восстановления (Колонка H)
                raw_recovery_date = sheet[f'H{row_idx}'].value
                if isinstance(raw_recovery_date, datetime):
                    recovery_date = raw_recovery_date.date()
                else:
                    recovery_date = datetime.strptime(str(raw_recovery_date).strip(), "%d.%m.%Y").date()

                # 2. Находим связанные объекты в базе Django
                machine = Machine.objects.filter(serial_number=serial_number).first()
                if not machine:
                    self.stdout.write(self.style.ERROR(f"Строка {row_idx}: Машина №{serial_number} не найдена в базе! Пропускаем."))
                    continue

                # Ищем или создаем автоматически элементы справочников на лету
                failure_node, _ = FailureNode.objects.get_or_create(name=node_name)
                recovery_method, _ = RecoveryMethod.objects.get_or_create(name=method_name)

                # 3. Сохраняем в базу данных Reclamation с вашими точными именами полей
                Reclamation.objects.update_or_create(
                    machine=machine,
                    refusal_date=refusal_date,
                    failure_node=failure_node,
                    defaults={
                        'operating_hours': operating_hours,
                        'failure_description': failure_description,
                        'recovery_method': recovery_method, # Передаем объект ForeignKey
                        'parts_used': spare_parts,          # Ваше точное имя поля из модели
                        'recovery_date': recovery_date,
                    }
                )
                self.stdout.write(self.style.SUCCESS(f"Строка {row_idx}: Рекламация по машине №{serial_number} успешно добавлена."))

            except Exception as e:
                self.stdout.write(self.style.ERROR(f"Ошибка в строке {row_idx}: {e}"))

        self.stdout.write(self.style.SUCCESS("Импорт рекламаций успешно завершен!"))
