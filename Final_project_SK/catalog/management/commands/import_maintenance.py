import os
from datetime import datetime
from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from catalog.models import Machine, Maintenance, MaintenanceType, ServiceCompany

class Command(BaseCommand):
    help = 'Импорт данных технического обслуживания из Excel data.xlsx (Вкладка ТО)'

    def handle(self, *args, **options):
        file_path = 'data.xlsx'
        
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f"Файл {file_path} не найден в корне проекта!"))
            return

        wb = load_workbook(file_path, data_only=True)
        # Открываем вкладку строго по её имени со скриншота
        sheet = wb['ТО output'] 

        self.stdout.write(self.style.WARNING("Начало импорта данных ТО..."))

        # Читаем строго со 2-й строки, так как шапка занимает всего 1 строку
        for row_idx in range(2, sheet.max_row + 1):
            serial_number_raw = sheet[f'A{row_idx}'].value # Зав. № машины в колонке A
            if not serial_number_raw:
                continue

            try:
                # Читаем ячейки строго по буквам колонок с вашего скриншота
                serial_number = str(serial_number_raw).strip()
                # Если номер импортировался как "3.0" или "17.0", убираем дробную часть
                if serial_number.endswith('.0'):
                    serial_number = serial_number[:-2]
                # Дописываем нули слева, чтобы совпало с базой (например, "0003")
                serial_number = serial_number.zfill(4)

                type_name = str(sheet[f'B{row_idx}'].value).strip()
                
                # Дата проведения ТО (Колонка C)
                raw_date = sheet[f'C{row_idx}'].value
                if isinstance(raw_date, datetime):
                    date = raw_date.date()
                else:
                    date = datetime.strptime(str(raw_date).strip(), "%d.%m.%Y").date()

                operating_hours = int(float(sheet[f'D{row_idx}'].value))
                order_number = str(sheet[f'E{row_idx}'].value).strip()
                
                # Дата заказ-наряда (Колонка F)
                raw_order_date = sheet[f'F{row_idx}'].value
                if isinstance(raw_order_date, datetime):
                    order_date = raw_order_date.date()
                else:
                    order_date = datetime.strptime(str(raw_order_date).strip(), "%d.%m.%Y").date()

                company_name = str(sheet[f'G{row_idx}'].value).strip()

                # 2. Ищем связанные объекты в базе данных Django
                machine = Machine.objects.filter(serial_number=serial_number).first()
                if not machine:
                    self.stdout.write(self.style.ERROR(f"Строка {row_idx}: Машина №{serial_number} не найдена в базе! Пропускаем."))
                    continue

                m_type = MaintenanceType.objects.filter(name__iexact=type_name).first()
                service_company = ServiceCompany.objects.filter(name__iexact=company_name).first()

                # 3. Записываем в базу данных
                Maintenance.objects.update_or_create(
                    order_number=order_number,
                    defaults={
                        'machine': machine,
                        'type': m_type,
                        'date': date,
                        'operating_hours': operating_hours,
                        'order_date': order_date,
                        'service_company': service_company,
                    }
                )
                self.stdout.write(self.style.SUCCESS(f"Строка {row_idx}: ТО для машины №{serial_number} успешно добавлено."))

            except Exception as e:
                self.stdout.write(self.style.ERROR(f"Ошибка в строке {row_idx}: {e}"))

        self.stdout.write(self.style.SUCCESS("Импорт ТО успешно завершен!"))
